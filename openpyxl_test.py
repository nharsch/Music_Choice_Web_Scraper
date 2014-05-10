from openpyxl import load_workbook
from openpyxl import Workbook

file_location = "csv_tester.xlsx"

wb2 = load_workbook(file_location)

ws1 = wb2.create_sheet()

wb2.save('new'+file_location)

print wb2.get_sheet_names()




